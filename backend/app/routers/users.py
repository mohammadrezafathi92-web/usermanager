import datetime as dt
import os
import tempfile
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy import func, nullsfirst, nullslast, or_
from sqlalchemy.orm import Session, joinedload

from .. import models, schemas
from ..database import get_db
from ..deps import get_current_admin
from ..services import user_ops

router = APIRouter(prefix="/api/users", tags=["users"], dependencies=[Depends(get_current_admin)])

# Same Persian labels as the panel's own frontend/src/utils.js STATUS_LABELS -
# kept in sync manually since the export below is generated server-side.
_STATUS_LABELS_FA = {
    "active": "فعال",
    "disabled": "غیرفعال",
    "quota_exceeded": "اتمام حجم",
    "expired": "منقضی‌شده",
}

# Columns the user list can be sorted by, from the "مرتب‌سازی" dropdown.
_SORT_COLUMNS = {
    "id": models.User.id,
    "username": models.User.username,
    "used_bytes": models.User.used_bytes,
    "total_quota_bytes": models.User.total_quota_bytes,
    "expire_at": models.User.expire_at,
    "status": models.User.status,
    "created_at": models.User.created_at,
}


def _charge_admin_for_package(db: Session, admin: models.AdminUser, package: models.Package, units: int = 1) -> None:
    """Atomically deducts `units` times the package's wholesale price (its
    cooperation_price, or the regular customer price if no cooperation
    price is configured) from a non-superadmin admin's own credit balance -
    what it costs them to provision this package for their own group.
    Superadmins own everything outright and are never charged. Uses a
    single conditional UPDATE (`WHERE balance >= cost`), same pattern as
    the customer wallet debit in routers/bot.py's add_balance, so two
    concurrent bulk-creates from the same admin can't both succeed past
    their real balance. Raises HTTPException(400) - and deducts nothing -
    if the balance can't cover it."""
    if admin.is_superadmin or units <= 0:
        return
    if admin.billing_mode == "usage":
        # این ادمین بابت هر پکیج پول کم نمی‌شود - اعتبارش به‌صورت حجمی
        # (volume_balance_gb) و لحظه‌ای در quota_manager.py's _apply_delta
        # کسر می‌شود، نه یکجا در لحظه ساخت کاربر.
        return
    unit_price = package.cooperation_price if package.cooperation_price is not None else (package.price or 0)
    cost = unit_price * units
    if cost <= 0:
        return
    result = db.execute(
        models.AdminUser.__table__.update()
        .where(models.AdminUser.id == admin.id, models.AdminUser.balance >= cost)
        .values(balance=models.AdminUser.balance - cost)
    )
    db.commit()
    if result.rowcount == 0:
        raise HTTPException(400, f"اعتبار شما کافی نیست - این پکیج {cost:,} تومان از اعتبار شما کم می‌کند")


def _refund_admin_for_package(db: Session, admin: models.AdminUser, package: models.Package, units: int) -> None:
    """Gives back credit reserved by _charge_admin_for_package for users
    that ended up NOT being created (e.g. bulk-create hit its collision
    safety cap before reaching the requested count) - see bulk_create_users
    below."""
    if admin.is_superadmin or units <= 0:
        return
    if admin.billing_mode == "usage":
        return
    unit_price = package.cooperation_price if package.cooperation_price is not None else (package.price or 0)
    amount = unit_price * units
    if amount <= 0:
        return
    db.execute(
        models.AdminUser.__table__.update()
        .where(models.AdminUser.id == admin.id)
        .values(balance=models.AdminUser.balance + amount)
    )
    db.commit()


def _get_owned_user(db: Session, admin: models.AdminUser, user_id: int) -> models.User:
    """Fetches a user and enforces group ownership: a non-superadmin gets a
    plain 404 (not 403 - deliberately doesn't confirm/deny whether the id
    belongs to someone else's group) for any user outside their own group.
    Superadmins bypass this entirely."""
    user = db.get(models.User, user_id)
    if not user:
        raise HTTPException(404, "کاربر پیدا نشد")
    if not admin.is_superadmin and user.owner_admin_id != admin.id:
        raise HTTPException(404, "کاربر پیدا نشد")
    return user


def _build_user_query(
    db: Session,
    admin: models.AdminUser,
    search: Optional[str] = None,
    status: Optional[models.UserStatus] = None,
    online_only: bool = False,
    owner_admin_id: Optional[int] = None,
    package_id: Optional[int] = None,
):
    """Shared filter-building for list_users and export_users below, so the
    Excel export always matches exactly what's on screen for the same
    search/status/online/group filters instead of drifting apart over
    time. Scoped to the caller's own group unless they're a superadmin -
    see list_users' docstring for the full rationale."""
    query = db.query(models.User)
    if not admin.is_superadmin:
        query = query.filter(models.User.owner_admin_id == admin.id)
    elif owner_admin_id is not None:
        query = query.filter(models.User.owner_admin_id == owner_admin_id)
    if search:
        like = f"%{search}%"
        query = query.filter(or_(models.User.username.ilike(like), models.User.full_name.ilike(like)))
    if status:
        query = query.filter(models.User.status == status)
    if package_id:
        query = query.filter(models.User.package_id == package_id)
    if online_only:
        # Users with at least one currently-open RADIUS session (openvpn/
        # l2tp) OR at least one xray connection the node last reported as
        # online (Connection.online, refreshed each poll_xray_node cycle -
        # see quota_manager.py).
        online_user_ids = (
            db.query(models.Connection.user_id)
            .outerjoin(models.RadiusActiveSession, models.RadiusActiveSession.connection_id == models.Connection.id)
            .filter(or_(models.RadiusActiveSession.id.isnot(None), models.Connection.online.is_(True)))
            .distinct()
        )
        query = query.filter(models.User.id.in_(online_user_ids))
    return query


@router.get("", response_model=schemas.UserListPage)
def list_users(
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
    page: int = 1,
    page_size: int = 50,
    search: Optional[str] = None,
    status: Optional[models.UserStatus] = None,
    online_only: bool = False,
    sort_by: str = "id",
    sort_dir: str = "desc",
    owner_admin_id: Optional[int] = None,
    package_id: Optional[int] = None,
):
    """Paginated + server-side-searched/filtered/sorted so the panel stays
    responsive with a large number of users (the old version loaded every
    user on every page load, which doesn't scale).

    Scoped to the caller's own group unless they're a superadmin - a
    superadmin sees everyone by default, or one specific group at a time
    via owner_admin_id (ignored for non-superadmins, who are always scoped
    to themselves regardless of what they pass)."""
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)

    query = _build_user_query(db, admin, search, status, online_only, owner_admin_id, package_id)
    total = query.with_entities(func.count(models.User.id)).scalar()

    sort_col = _SORT_COLUMNS.get(sort_by, models.User.id)
    order_by_cols = []
    if sort_by == "expire_at":
        # NULL expire_at means "never expires" - semantically that's
        # INFINITELY far in the future, not the smallest possible value
        # (SQLite's native NULL-sorting rule, which otherwise dumps these
        # rows at the very bottom even when sorting "most time left first" /
        # desc - the opposite of what an admin scanning for the
        # least-urgent accounts would expect, and what made the whole
        # column feel "random" - see nullsfirst/nullslast below).
        order_by_cols.append(nullsfirst(sort_col.desc()) if sort_dir != "asc" else nullslast(sort_col.asc()))
        # Two very different "no date" cases both have expire_at=NULL -
        # truly-unlimited users AND "counts from first use" users still
        # awaiting activation (see UserListItem.expire_days_after_first_use).
        # Without a secondary key these interleave in arbitrary/insertion
        # order among themselves, which looks erratic even though the
        # real-dated rows above/below them are correctly sorted - group the
        # pending-activation ones together instead of scattering them.
        order_by_cols.append(models.User.expire_days_after_first_use.isnot(None).desc())
    else:
        order_by_cols.append(sort_col.asc() if sort_dir == "asc" else sort_col.desc())
    order_by_cols.append(models.User.id.desc())  # stable tiebreaker for equal/NULL values

    users = (
        query.options(joinedload(models.User.connections), joinedload(models.User.owner_admin))
        .order_by(*order_by_cols)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    # Which of these users have a currently-open RADIUS (openvpn/l2tp)
    # session - xray's online flag is already sitting on each Connection
    # row (joinedload'ed above), so no extra query is needed for that part.
    user_ids = [u.id for u in users]
    ppp_online_ids = set()
    if user_ids:
        ppp_online_ids = {
            uid
            for (uid,) in db.query(models.Connection.user_id)
            .join(models.RadiusActiveSession, models.RadiusActiveSession.connection_id == models.Connection.id)
            .filter(models.Connection.user_id.in_(user_ids))
            .distinct()
        }

    items = []
    for u in users:
        item = schemas.UserListItem.model_validate(u)
        item.connections_count = len(u.connections)
        item.online = u.id in ppp_online_ids or any(c.online for c in u.connections)
        items.append(item)

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/ids", response_model=list[int])
def list_user_ids(
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
    search: Optional[str] = None,
    status: Optional[models.UserStatus] = None,
    online_only: bool = False,
    owner_admin_id: Optional[int] = None,
    package_id: Optional[int] = None,
):
    """Returns EVERY user id matching these filters, ignoring pagination -
    powers the "انتخاب همه با این فیلتر" button in Users.jsx so a group
    action (e.g. disable/renew/change volume for a package) can target every
    matching user across all pages, not just the 50 currently on screen.
    Registered BEFORE GET /{user_id} below for the same reason as /export.
    Capped at 5000 ids per call as a sanity limit for one bulk action."""
    query = _build_user_query(db, admin, search, status, online_only, owner_admin_id, package_id)
    return [uid for (uid,) in query.with_entities(models.User.id).limit(5000).all()]


@router.post("/bulk", response_model=schemas.BulkCreateUsersResult)
def bulk_create_users(
    payload: schemas.BulkCreateUsersRequest,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    # Validated here, BEFORE any wallet charge below - user_ops.bulk_create_users
    # re-checks this too, but doing it only there meant an invalid count
    # (e.g. 0 or >1000) let the charge below commit first and then throw,
    # permanently debiting the admin's wallet for zero created users since
    # the refund line is never reached on an exception. Keep both checks:
    # this one guards the charge, user_ops' own guards it against being
    # called with a bad count directly (bot API, future callers, ...).
    if payload.count <= 0:
        raise HTTPException(400, "تعداد باید بزرگتر از صفر باشد")
    if payload.count > 1000:
        raise HTTPException(400, "حداکثر ۱۰۰۰ کاربر در هر بار")

    # Non-superadmins must always build from a package - enforced here too
    # (not just hidden in the UI - see Users.jsx) since a direct API call
    # could otherwise skip the cooperation-price wallet charge entirely by
    # never going through _charge_admin_for_package below.
    if not admin.is_superadmin and not payload.package_id:
        raise HTTPException(400, "ساخت کاربر بدون پکیج مجاز نیست - یک پکیج انتخاب کنید")

    package = None
    if payload.package_id:
        package = db.get(models.Package, payload.package_id)
        if not package:
            raise HTTPException(400, "پکیج پیدا نشد")
        # Reserve the worst case (every one of `count` succeeds) upfront -
        # refunded below for any that don't actually get created (fully, if
        # user_ops.bulk_create_users itself raises partway through).
        _charge_admin_for_package(db, admin, package, units=payload.count)

    try:
        result = user_ops.bulk_create_users(
            db,
            prefix=payload.prefix,
            count=payload.count,
            package_id=payload.package_id,
            quota_gb=payload.quota_gb,
            expire_days=payload.expire_days,
            notes=payload.notes,
            connections=payload.connections,
            owner_admin_id=admin.id,
        )
    except Exception:
        if package:
            _refund_admin_for_package(db, admin, package, units=payload.count)
        raise
    if package and result["created_count"] < payload.count:
        _refund_admin_for_package(db, admin, package, units=payload.count - result["created_count"])
    return result


@router.put("/bulk", response_model=schemas.BulkUpdateUsersResult)
def bulk_update_users(
    payload: schemas.BulkUpdateUsersRequest,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    package = None
    if payload.package_id:
        package = db.get(models.Package, payload.package_id)
        if not package:
            raise HTTPException(400, "پکیج پیدا نشد")

    return user_ops.bulk_update_users(
        db,
        user_ids=payload.user_ids,
        add_gb=payload.add_gb,
        add_days=payload.add_days,
        reset_usage=payload.reset_usage,
        status=payload.status,
        max_concurrent_sessions=payload.max_concurrent_sessions,
        package=package,
        owner_admin_id=None if admin.is_superadmin else admin.id,
    )


@router.delete("/bulk", response_model=schemas.BulkDeleteUsersResult)
def bulk_delete_users(
    payload: schemas.BulkDeleteUsersRequest,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    return user_ops.bulk_delete_users(
        db, user_ids=payload.user_ids, owner_admin_id=None if admin.is_superadmin else admin.id
    )


@router.post("", response_model=schemas.UserOut)
def create_user(
    payload: schemas.UserCreate,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    if db.query(models.User).filter(models.User.username == payload.username).first():
        raise HTTPException(400, "این نام کاربری قبلا ثبت شده است")

    # Non-superadmins must always build from a package - enforced here too
    # (not just hidden in the UI - see Users.jsx) since a direct API call
    # could otherwise skip the cooperation-price wallet charge entirely.
    if not admin.is_superadmin and not payload.package_id:
        raise HTTPException(400, "ساخت کاربر بدون پکیج مجاز نیست - یک پکیج انتخاب کنید")

    # Only a superadmin may hand a new user to a DIFFERENT admin's group -
    # everyone else's users always land in their own group, regardless of
    # what owner_admin_id they send.
    owner_admin_id = admin.id
    if admin.is_superadmin and payload.owner_admin_id is not None:
        if not db.get(models.AdminUser, payload.owner_admin_id):
            raise HTTPException(400, "ادمین مقصد پیدا نشد")
        owner_admin_id = payload.owner_admin_id

    data = payload.model_dump(exclude={"package_id", "owner_admin_id"})
    data["owner_admin_id"] = owner_admin_id
    package = None
    if payload.package_id:
        package = db.get(models.Package, payload.package_id)
        if not package:
            raise HTTPException(400, "پکیج پیدا نشد")
        _charge_admin_for_package(db, admin, package, units=1)
        # the package's own quota/duration/concurrent-session cap win over
        # whatever was in the manual fields above
        data["total_quota_bytes"] = int(package.quota_gb * 1024 ** 3) if package.quota_gb else 0
        data["expire_at"] = (
            dt.datetime.utcnow() + dt.timedelta(days=package.duration_days) if package.duration_days else None
        )
        data["expire_days_after_first_use"] = None
        data["max_concurrent_sessions"] = package.max_concurrent_sessions
        data["package_id"] = package.id

    user = models.User(**data)
    db.add(user)
    db.commit()
    db.refresh(user)

    if package:
        user_ops.provision_package_connections(db, user, package)
        db.commit()
        db.refresh(user)

    return user


@router.get("/export")
def export_users(
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
    search: Optional[str] = None,
    status: Optional[models.UserStatus] = None,
    online_only: bool = False,
    owner_admin_id: Optional[int] = None,
    package_id: Optional[int] = None,
):
    """Exports the same set of users list_users would show for these exact
    filters (search/status/online/group/package) as a formatted .xlsx file -
    "اکسپورت گروهی" in Users.jsx. Registered BEFORE GET /{user_id} below so
    the literal path "/export" is matched first (otherwise Starlette would
    try to parse "export" as an int user_id and 422)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    query = _build_user_query(db, admin, search, status, online_only, owner_admin_id, package_id)
    users = (
        query.options(joinedload(models.User.owner_admin))
        .order_by(models.User.id.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "کاربران"
    ws.sheet_view.rightToLeft = True

    headers = [
        "شناسه", "نام کاربری", "نام کامل", "وضعیت", "مصرف (GB)", "حجم مجاز (GB)",
        "تاریخ انقضا", "موجودی (تومان)", "تعداد سرویس", "ادمین مالک", "تاریخ ثبت",
    ]
    header_fill = PatternFill(start_color="4763F5", end_color="4763F5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for row, u in enumerate(users, start=2):
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.username)
        ws.cell(row=row, column=3, value=u.full_name or "")
        ws.cell(row=row, column=4, value=_STATUS_LABELS_FA.get(u.status.value if hasattr(u.status, "value") else u.status, str(u.status)))
        ws.cell(row=row, column=5, value=round((u.used_bytes or 0) / 1024 ** 3, 2))
        ws.cell(row=row, column=6, value=round((u.total_quota_bytes or 0) / 1024 ** 3, 2) if u.total_quota_bytes else "نامحدود")
        ws.cell(row=row, column=7, value=u.expire_at.strftime("%Y-%m-%d %H:%M") if u.expire_at else "بدون انقضا")
        ws.cell(row=row, column=8, value=u.balance or 0)
        ws.cell(row=row, column=9, value=len(u.connections))
        ws.cell(row=row, column=10, value=u.owner_admin.username if u.owner_admin else "")
        ws.cell(row=row, column=11, value=u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    widths = [8, 20, 20, 12, 12, 14, 18, 14, 12, 16, 18]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(tmp_path)

    stamp = dt.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return FileResponse(
        tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"users_export_{stamp}.xlsx",
        background=None,
    )


@router.get("/{user_id}", response_model=schemas.UserOut)
def get_user(user_id: int, db: Session = Depends(get_db), admin: models.AdminUser = Depends(get_current_admin)):
    user = _get_owned_user(db, admin, user_id)
    out = schemas.UserOut.model_validate(user)
    if out.connections:
        counts = dict(
            db.query(models.RadiusActiveSession.connection_id, func.count(models.RadiusActiveSession.id))
            .filter(models.RadiusActiveSession.connection_id.in_([c.id for c in out.connections]))
            .group_by(models.RadiusActiveSession.connection_id)
            .all()
        )
        for c in out.connections:
            c.active_session_count = counts.get(c.id, 0)
            if c.type in (models.ConnectionType.openvpn, models.ConnectionType.l2tp, models.ConnectionType.ikev2, models.ConnectionType.sstp):
                # PPP connections have no "online" DB column of their own
                # (that column only means anything for xray) - their live
                # state is whether a RADIUS session is currently open.
                c.online = c.active_session_count > 0
    return out


@router.put("/{user_id}", response_model=schemas.UserOut)
def update_user(
    user_id: int,
    payload: schemas.UserUpdate,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    user = _get_owned_user(db, admin, user_id)

    data = payload.model_dump(exclude_unset=True)

    # Not a real column on User - handled specially below.
    clear_trigger = data.pop("clear_expire_days_trigger", None)

    if "owner_admin_id" in data:
        if not admin.is_superadmin:
            # Non-superadmins can't reassign a user to another group at
            # all - silently drop the field rather than error, so the same
            # request body works regardless of who sends it.
            data.pop("owner_admin_id")
        elif data["owner_admin_id"] is not None and not db.get(models.AdminUser, data["owner_admin_id"]):
            raise HTTPException(400, "ادمین مقصد پیدا نشد")

    # NOTE: telegram_id is intentionally allowed on more than one User - a
    # single Telegram account can have several separate panel accounts (a
    # customer who bought more than once ended up with more than one
    # username). The bot shows an account picker in that case - see
    # routers/bot.py's list_users_by_telegram and telegram_bot's
    # _resolve_account.

    if data.get("expire_at") is not None:
        # A fixed expiry date was explicitly set - it takes precedence over
        # any pending "count from first use" plan.
        user.expire_days_after_first_use = None
    if data.get("expire_days_after_first_use"):
        # Switching to "count from first use" - clear any fixed date so the
        # RADIUS auth handler knows expire_at still needs to be computed.
        data["expire_at"] = None
    if clear_trigger:
        user.expire_days_after_first_use = None

    # max_concurrent_sessions is a real column on User now (a combined cap
    # across all of their connections - see models.py) so it just flows
    # through the generic setattr loop below like everything else.
    for k, v in data.items():
        setattr(user, k, v)

    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/reset-usage", response_model=schemas.UserOut)
def reset_usage(user_id: int, db: Session = Depends(get_db), admin: models.AdminUser = Depends(get_current_admin)):
    user = _get_owned_user(db, admin, user_id)
    user.used_bytes = 0
    if user.status == models.UserStatus.quota_exceeded:
        user.status = models.UserStatus.active
    db.commit()
    db.refresh(user)
    return user


@router.delete("/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), admin: models.AdminUser = Depends(get_current_admin)):
    user = _get_owned_user(db, admin, user_id)
    user_ops.delete_user_cascade(db, user)
    return {"ok": True}


# ---------------------------------------------------------------- connections
def _get_user_and_node(
    db: Session, admin: models.AdminUser, user_id: int, node_id: int
) -> tuple[models.User, models.Node]:
    user = _get_owned_user(db, admin, user_id)
    node = db.get(models.Node, node_id)
    if not node:
        raise HTTPException(400, "نود پیدا نشد")
    return user, node


@router.post("/{user_id}/connections/wireguard", response_model=schemas.ConnectionOut)
def add_wireguard_connection(
    user_id: int,
    payload: schemas.ConnectionCreateWireguard,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    user, node = _get_user_and_node(db, admin, user_id, payload.node_id)
    return user_ops.provision_wireguard(db, user, node)


@router.post("/{user_id}/connections/openvpn", response_model=schemas.ConnectionOut)
def add_openvpn_connection(
    user_id: int,
    payload: schemas.ConnectionCreateOpenvpn,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    user, node = _get_user_and_node(db, admin, user_id, payload.node_id)
    return user_ops.provision_openvpn(db, user, node, payload.max_concurrent_sessions)


@router.post("/{user_id}/connections/l2tp", response_model=schemas.ConnectionOut)
def add_l2tp_connection(
    user_id: int,
    payload: schemas.ConnectionCreateL2tp,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    user, node = _get_user_and_node(db, admin, user_id, payload.node_id)
    return user_ops.provision_l2tp(db, user, node, payload.max_concurrent_sessions)


@router.post("/{user_id}/connections/ikev2", response_model=schemas.ConnectionOut)
def add_ikev2_connection(
    user_id: int,
    payload: schemas.ConnectionCreateIkev2,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    user, node = _get_user_and_node(db, admin, user_id, payload.node_id)
    return user_ops.provision_ikev2(db, user, node, payload.max_concurrent_sessions)


@router.post("/{user_id}/connections/sstp", response_model=schemas.ConnectionOut)
def add_sstp_connection(
    user_id: int,
    payload: schemas.ConnectionCreateSstp,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    user, node = _get_user_and_node(db, admin, user_id, payload.node_id)
    return user_ops.provision_sstp(db, user, node, payload.max_concurrent_sessions)


@router.post("/{user_id}/connections/xray", response_model=schemas.ConnectionOut)
def add_xray_connection(
    user_id: int,
    payload: schemas.ConnectionCreateXray,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    user, node = _get_user_and_node(db, admin, user_id, payload.node_id)
    return user_ops.provision_xray(db, user, node, payload.flow or "")


@router.put("/{user_id}/connections/{connection_id}", response_model=schemas.ConnectionOut)
def update_connection(
    user_id: int,
    connection_id: int,
    payload: schemas.ConnectionUpdate,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    """Lets the admin change a connection's simultaneous-session limit,
    manually enable/disable it, or clear a temporary ban (set banned_until
    to null or a past date)."""
    _get_owned_user(db, admin, user_id)
    conn = db.get(models.Connection, connection_id)
    if not conn or conn.user_id != user_id:
        raise HTTPException(404, "کانکشن پیدا نشد")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(conn, k, v)
    db.commit()
    db.refresh(conn)
    out = schemas.ConnectionOut.model_validate(conn)
    out.active_session_count = db.query(models.RadiusActiveSession).filter(
        models.RadiusActiveSession.connection_id == conn.id
    ).count()
    if conn.type in (models.ConnectionType.openvpn, models.ConnectionType.l2tp, models.ConnectionType.ikev2, models.ConnectionType.sstp):
        out.online = out.active_session_count > 0
    return out


@router.delete("/{user_id}/connections/{connection_id}")
def delete_connection(
    user_id: int,
    connection_id: int,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    _get_owned_user(db, admin, user_id)
    conn = db.get(models.Connection, connection_id)
    if not conn or conn.user_id != user_id:
        raise HTTPException(404, "کانکشن پیدا نشد")
    user_ops.delete_connection(db, conn)
    return {"ok": True}


@router.get("/{user_id}/connections/{connection_id}/share", response_model=schemas.ConnectionShareLink)
def get_share_link(
    user_id: int,
    connection_id: int,
    db: Session = Depends(get_db),
    admin: models.AdminUser = Depends(get_current_admin),
):
    _get_owned_user(db, admin, user_id)
    conn = db.get(models.Connection, connection_id)
    if not conn or conn.user_id != user_id:
        raise HTTPException(404, "کانکشن پیدا نشد")
    share = user_ops.get_connection_share(conn)
    return schemas.ConnectionShareLink(connection_id=conn.id, **share)
