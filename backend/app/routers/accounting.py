"""The "حساب‌داری" (Accounting) section's API - thin role-scoped reads over
services/accounting.py's ledger plus superadmin-only manual expense entry.
Design agreed with the panel owner 2026-08-08 (docs/accounting-design.md):
superadmin sees the whole panel (incl. expenses + net profit), a level-2
Admin sees their own tree (self + their sellers), a Seller only themselves -
all enforced by accounting.scoped_query(), never by the frontend."""
import datetime as dt
import os
import tempfile
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from .. import models, schemas
from ..database import get_db
from ..services.jalali import fmt_jalali
from ..deps import get_current_admin, require_superadmin, require_confirm_password
from ..services import accounting, hierarchy

router = APIRouter(prefix="/api/accounting", tags=["accounting"])


def _parse_date(value: Optional[str], end: bool = False) -> Optional[dt.datetime]:
    """YYYY-MM-DD -> datetime; `end` dates become exclusive midnight-after
    so a single-day range [d, d] covers that whole day."""
    if not value:
        return None
    try:
        parsed = dt.datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, "فرمت تاریخ باید YYYY-MM-DD باشد")
    return parsed + dt.timedelta(days=1) if end else parsed


@router.get("/summary")
def get_summary(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current: models.AdminUser = Depends(get_current_admin),
):
    out = accounting.summary(
        db, current,
        date_from=_parse_date(date_from), date_to=_parse_date(date_to, end=True),
    )
    out["role"] = hierarchy.role(current)
    return out


@router.get("/subtree")
def get_subtree_rollup(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current: models.AdminUser = Depends(get_current_admin),
):
    """One row per direct sub-account - see accounting.subtree_rollup.

    Returns an empty list rather than 403 for a Seller, who simply has no
    sub-accounts. A 403 would say "you are not allowed", which is not true
    and would make the frontend show an error where there is only nothing
    to show.
    """
    return accounting.subtree_rollup(
        db, current,
        date_from=_parse_date(date_from), date_to=_parse_date(date_to, end=True),
    )


@router.get("/series")
def get_series(
    granularity: str = "day",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current: models.AdminUser = Depends(get_current_admin),
):
    if granularity not in ("day", "month"):
        raise HTTPException(400, "granularity باید day یا month باشد")
    return accounting.series(
        db, current, granularity,
        date_from=_parse_date(date_from), date_to=_parse_date(date_to, end=True),
    )


@router.get("/transactions", response_model=schemas.LedgerPage)
def list_transactions(
    page: int = 1,
    page_size: int = 50,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    kind: Optional[str] = None,
    admin_id: Optional[int] = None,
    payment_card_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current: models.AdminUser = Depends(get_current_admin),
):
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)
    q = accounting.apply_filters(
        accounting.scoped_query(db, current),
        date_from=_parse_date(date_from), date_to=_parse_date(date_to, end=True),
        kind=kind, admin_id=admin_id, payment_card_id=payment_card_id,
    )
    total = q.count()
    items = (
        q.order_by(models.LedgerEntry.created_at.desc(), models.LedgerEntry.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return schemas.LedgerPage(
        items=[schemas.LedgerEntryOut.model_validate(e) for e in items],
        total=total, page=page, page_size=page_size,
    )


@router.post("/expenses", response_model=schemas.LedgerEntryOut)
def create_expense(
    payload: schemas.ExpenseCreate,
    db: Session = Depends(get_db),
    current: models.AdminUser = Depends(require_superadmin),
):
    if payload.amount <= 0:
        raise HTTPException(400, "مبلغ هزینه باید بزرگ‌تر از صفر باشد")
    entry = accounting.record(
        db, "expense", payload.amount,
        actor_admin_id=current.id,
        category=(payload.category or "").strip() or None,
        note=(payload.note or "").strip() or None,
        created_at=payload.created_at,
    )
    db.commit()
    db.refresh(entry)
    return schemas.LedgerEntryOut.model_validate(entry)


@router.delete("/expenses/{entry_id}")
def delete_expense(
    entry_id: int,
    db: Session = Depends(get_db),
    current: models.AdminUser = Depends(require_superadmin), _confirm=Depends(require_confirm_password)):
    """Only manual expense rows are deletable - automatic sale/credit rows
    are the books themselves and stay immutable."""
    entry = db.get(models.LedgerEntry, entry_id)
    if not entry or entry.kind != "expense":
        raise HTTPException(404, "هزینه پیدا نشد")
    db.delete(entry)
    db.commit()
    return {"ok": True}


@router.get("/export")
def export_xlsx(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    kind: Optional[str] = None,
    db: Session = Depends(get_db),
    current: models.AdminUser = Depends(get_current_admin),
):
    """Excel export of the (role-scoped, filtered) transactions - same
    openpyxl+FileResponse approach as routers/users.py's export."""
    from openpyxl import Workbook

    q = accounting.apply_filters(
        accounting.scoped_query(db, current),
        date_from=_parse_date(date_from), date_to=_parse_date(date_to, end=True),
        kind=kind,
    ).order_by(models.LedgerEntry.created_at.desc())

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = [
        "ID", "Kind", "Amount (Toman)", "Customer", "Admin/Seller",
        "Package", "Payment method", "Card ID", "Discount code",
        "Discount amount", "Category", "Note", "Created at",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row, e in enumerate(q.all(), start=2):
        ws.cell(row=row, column=1, value=e.id)
        ws.cell(row=row, column=2, value=e.kind)
        ws.cell(row=row, column=3, value=e.amount)
        ws.cell(row=row, column=4, value=e.username_snapshot)
        ws.cell(row=row, column=5, value=e.admin_username_snapshot)
        ws.cell(row=row, column=6, value=e.package_name_snapshot)
        ws.cell(row=row, column=7, value=e.payment_method)
        ws.cell(row=row, column=8, value=e.payment_card_id)
        ws.cell(row=row, column=9, value=e.discount_code)
        ws.cell(row=row, column=10, value=e.discount_amount)
        ws.cell(row=row, column=11, value=e.category)
        ws.cell(row=row, column=12, value=e.note)
        ws.cell(row=row, column=13, value=fmt_jalali(e.created_at) if e.created_at else None)

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    filename = f"accounting-{dt.datetime.utcnow().strftime('%Y%m%d-%H%M')}.xlsx"
    return FileResponse(path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
